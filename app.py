from flask import Flask, jsonify, request, render_template, send_file, session, redirect
import sqlite3, os, io, hashlib, secrets, time
from datetime import datetime, date, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)


def _secret_key():
    """Clé de session persistée localement (hors git) : les sessions survivent aux redémarrages."""
    path = os.path.join(os.path.dirname(__file__), "secret_key.txt")
    try:
        k = open(path).read().strip()
        if k:
            return k
    except Exception:
        pass
    k = secrets.token_hex(32)
    try:
        open(path, "w").write(k)
    except Exception:
        pass
    return k


app.secret_key = _secret_key()
app.config["TEMPLATES_AUTO_RELOAD"] = True  # recharge les templates modifiés sans redémarrer
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)  # expiration de session
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

DB = os.path.join(os.path.dirname(__file__), "epie_crm.db")


def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def hash_pw(pw):
    """Ancien hachage (SHA-256 non salé) — conservé pour vérifier les comptes existants."""
    return hashlib.sha256(pw.encode()).hexdigest()


def make_pw_hash(pw):
    """Hachage renforcé : PBKDF2-SHA256 salé (fourni par Werkzeug, inclus dans Flask)."""
    return generate_password_hash(pw)


def verify_pw(stored, pw):
    """Vérifie un mot de passe quel que soit le format du hash stocké (ancien ou nouveau)."""
    if not stored:
        return False
    if stored.startswith(("pbkdf2:", "scrypt:")):
        try:
            return check_password_hash(stored, pw)
        except Exception:
            return False
    return stored == hash_pw(pw)  # ancien format SHA-256


# ── Limitation des tentatives de connexion ──────────────────────────────────
LOGIN_MAX_ATTEMPTS = 5     # échecs avant blocage
LOGIN_WINDOW = 15 * 60     # fenêtre de comptage / durée du blocage (15 min)
_login_attempts = {}       # "ip|login" -> [nb_échecs, timestamp_premier_échec]


def _login_check(key):
    """Minutes de blocage restantes, ou 0 si la connexion est autorisée."""
    rec = _login_attempts.get(key)
    if not rec:
        return 0
    count, first = rec
    if time.time() - first > LOGIN_WINDOW:
        _login_attempts.pop(key, None)
        return 0
    if count >= LOGIN_MAX_ATTEMPTS:
        return max(1, int((LOGIN_WINDOW - (time.time() - first)) // 60))
    return 0


def _login_fail(key):
    rec = _login_attempts.get(key)
    if rec and time.time() - rec[1] <= LOGIN_WINDOW:
        rec[0] += 1
    else:
        _login_attempts[key] = [1, time.time()]
    return _login_attempts[key][0]


# ── Journal d'audit ─────────────────────────────────────────────────────────
def audit(action, details="", login=None):
    """Trace une action sensible. Ne bloque jamais l'application en cas d'erreur."""
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO audit_log (user_login, role, action, details) VALUES (?,?,?,?)",
            (login or session.get("login", "?"), session.get("role", ""), action, str(details)[:500])
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


# ── Init DB ──────────────────────────────────────────────────────────────────


def _anciennete(debut):
    """Calcule l'ancienneté à partir de la date d'entrée."""
    if not debut:
        return ""
    try:
        d = datetime.strptime(debut, "%Y-%m-%d")
        delta = date.today() - d.date()
        years = delta.days // 365
        months = (delta.days % 365) // 30
        if years >= 1:
            return f"{years} an{'s' if years > 1 else ''}{' '+str(months)+'m' if months else ''}"
        return f"{months} mois" if months else "< 1 mois"
    except Exception:
        return ""


def _age(dn):
    """Calcule l'âge à partir de la date de naissance."""
    if not dn:
        return None
    try:
        d = datetime.strptime(dn, "%Y-%m-%d")
        today = date.today()
        return today.year - d.year - ((today.month, today.day) < (d.month, d.day))
    except Exception:
        return None

# ── Sauvegarde de la base ───────────────────────────────────────────────────
BACKUP_DIR = os.path.join(os.path.dirname(__file__), "backups")
BACKUP_KEEP = 30  # nombre de copies conservées


def backup_db():
    """Copie horodatée de la base dans backups/ via l'API SQLite (fiable à chaud).
    Rotation automatique : seules les BACKUP_KEEP copies les plus récentes sont gardées."""
    if not os.path.exists(DB):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    name = f"epie_crm_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    try:
        src = sqlite3.connect(DB)
        dst = sqlite3.connect(os.path.join(BACKUP_DIR, name))
        src.backup(dst)
        dst.close()
        src.close()
    except Exception:
        return None
    # Rotation
    files = sorted(
        f for f in os.listdir(BACKUP_DIR)
        if f.startswith("epie_crm_") and f.endswith(".db")
    )
    for old in files[:-BACKUP_KEEP]:
        try:
            os.remove(os.path.join(BACKUP_DIR, old))
        except Exception:
            pass
    return name


# ── Calendrier : jours ouvrés & jours fériés français ──────────────────────

def _paques(year):
    """Dimanche de Pâques (algorithme de Butcher, valable toute année grégorienne)."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mois, jour = divmod(h + l - 7 * m + 114, 31)
    return date(year, mois, jour + 1)


def jours_feries(year):
    """Les 11 jours fériés français (métropole) pour une année donnée."""
    p = _paques(year)
    return {
        date(year, 1, 1),    # Jour de l'an
        p + timedelta(days=1),   # Lundi de Pâques
        date(year, 5, 1),    # Fête du travail
        date(year, 5, 8),    # Victoire 1945
        p + timedelta(days=39),  # Ascension
        p + timedelta(days=50),  # Lundi de Pentecôte
        date(year, 7, 14),   # Fête nationale
        date(year, 8, 15),   # Assomption
        date(year, 11, 1),   # Toussaint
        date(year, 11, 11),  # Armistice 1918
        date(year, 12, 25),  # Noël
    }


def jours_ouvres(debut, fin, creneau="journee"):
    """Nombre de jours ouvrés (lun-ven, hors fériés) entre deux dates incluses.
    Demi-journée (matin/apm) : 0.5 si le jour est ouvré, sinon 0."""
    try:
        d0 = datetime.strptime(str(debut)[:10], "%Y-%m-%d").date()
        d1 = datetime.strptime(str(fin)[:10], "%Y-%m-%d").date()
    except Exception:
        return 0
    if d1 < d0:
        return 0
    feries = set()
    for y in range(d0.year, d1.year + 1):
        feries |= jours_feries(y)
    n = 0
    d = d0
    while d <= d1:
        if d.weekday() < 5 and d not in feries:
            n += 1
        d += timedelta(days=1)
    if creneau in ("matin", "apm"):
        return 0.5 if n > 0 else 0
    return n


# Seuls ces types de congé décomptent le solde de congés payés
TYPES_DECOMPTES = ("Congés payés",)


def _conge_row(row):
    """Enrichit une ligne congé avec le nombre de jours ouvrés calculé serveur."""
    d = dict(row)
    d["jours"] = jours_ouvres(d.get("debut"), d.get("fin"), d.get("creneau") or "journee")
    return d


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.executescript("""
    CREATE TABLE IF NOT EXISTS employees (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      prenom TEXT NOT NULL,
      nom TEXT NOT NULL,
      poste TEXT,
      contrat TEXT CHECK(contrat IN ('CDI','CDD','Alternant','Autoentrepreneur','Stagiaire','Contrat pro')),
      email TEXT,
      tel TEXT,
      debut TEXT,
      date_fin TEXT,
      solde_cp INTEGER DEFAULT 25,
      actif INTEGER DEFAULT 1,
      created_at TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS users (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      login TEXT UNIQUE NOT NULL,
      password TEXT NOT NULL,
      role TEXT CHECK(role IN ('directeur','rh','employe')) DEFAULT 'employe',
      emp_id INTEGER REFERENCES employees(id),
      created_at TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS conges (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      emp_id INTEGER NOT NULL REFERENCES employees(id),
      type TEXT,
      debut TEXT NOT NULL,
      fin TEXT NOT NULL,
      creneau TEXT DEFAULT 'journee',
      motif TEXT,
      statut TEXT DEFAULT 'En attente' CHECK(statut IN ('En attente','Approuvé','Refusé')),
      note_refus TEXT DEFAULT '',
      created_at TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS absences (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      emp_id INTEGER NOT NULL REFERENCES employees(id),
      date TEXT NOT NULL,
      type TEXT DEFAULT 'Absence' CHECK(type IN ('Absence','Retard')),
      justifiee INTEGER DEFAULT 0,
      note TEXT,
      created_at TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS audit_log (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      ts TEXT DEFAULT (datetime('now','localtime')),
      user_login TEXT,
      role TEXT,
      action TEXT,
      details TEXT
    );
    """)

    # Ajout colonne duree sur absences
    try:
        c.execute("ALTER TABLE absences ADD COLUMN duree INTEGER DEFAULT 0")
    except Exception:
        pass

    # Ajout colonne must_change_pw sur users
    try:
        c.execute("ALTER TABLE users ADD COLUMN must_change_pw INTEGER DEFAULT 0")
    except Exception:
        pass

    # Ajout colonnes manquantes (compatibilité upgrade)
    for col_sql in [
        "ALTER TABLE employees ADD COLUMN date_naissance TEXT",
        "ALTER TABLE employees ADD COLUMN tel_urgence TEXT DEFAULT ''",
        "ALTER TABLE conges ADD COLUMN heures TEXT DEFAULT ''",
    ]:
        try:
            c.execute(col_sql)
        except Exception:
            pass

    # Données de démo si vide
    if c.execute("SELECT COUNT(*) FROM employees").fetchone()[0] == 0:
        emps = [
            ("Amina","Benali","Directrice","CDI","a.benali@epie-formation.fr","06 12 34 56 78","2018-03-01",None,25),
            ("Thomas","Girard","Formateur","CDI","t.girard@epie-formation.fr","06 23 45 67 89","2019-09-01",None,25),
            ("Lucie","Martin","Assistante pédagogique","CDI","l.martin@epie-formation.fr","06 34 56 78 90","2020-01-15",None,25),
            ("Karim","Ouali","Formateur","CDD","k.ouali@epie-formation.fr","06 45 67 89 01","2025-09-01","2026-08-31",0),
            ("Sonia","Petit","Comptable","CDI","s.petit@epie-formation.fr","06 56 78 90 12","2021-06-01",None,25),
            ("Cyl","Meziane","Alternante Data/IA","Alternant","c.meziane@epie-formation.fr","06 67 89 01 23","2024-09-02","2026-08-31",0),
            ("Jean-Pierre","Rossi","Intervenant","Autoentrepreneur","jp.rossi@gmail.com","06 78 90 12 34","2023-01-10",None,0),
            ("Fatima","Zerrouk","Formatrice","CDI","f.zerrouk@epie-formation.fr","06 89 01 23 45","2022-04-01",None,25),
            ("Mathieu","Durand","Chargé de communication","CDD","m.durand@epie-formation.fr","06 90 12 34 56","2025-03-01","2026-09-15",0),
            ("Nadia","Brahim","Intervenante","Autoentrepreneur","n.brahim@gmail.com","07 01 23 45 67","2024-02-01",None,0),
        ]
        c.executemany("""
            INSERT INTO employees (prenom,nom,poste,contrat,email,tel,debut,date_fin,solde_cp)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, emps)

        users_data = [
            ("directeur", hash_pw("epie2024"), "directeur", 1),
            ("rh",        hash_pw("rh2024"),   "rh",        5),
        ]
        c.executemany("""
            INSERT INTO users (login,password,role,emp_id,must_change_pw)
            VALUES (?,?,?,?,1)
        """, users_data)

        # Comptes employés auto (login: initiale prénom + nom, mdp: epie123)
        for row in c.execute("SELECT id,prenom,nom FROM employees").fetchall():
            login = (row["prenom"][0] + row["nom"]).lower().replace(" ","").replace("-","")
            c.execute("""
                INSERT OR IGNORE INTO users (login,password,role,emp_id,must_change_pw)
                VALUES (?,?,?,?,1)
            """, (login, hash_pw("epie123"), "employe", row["id"]))

        c.executemany("""
            INSERT INTO conges (emp_id,type,debut,fin,motif,statut,note_refus)
            VALUES (?,?,?,?,?,?,?)
        """, [
            (6,"Congés payés","2026-04-07","2026-04-11","Vacances Pâques","En attente",""),
            (2,"RTT","2026-03-28","2026-03-28","RTT","Approuvé",""),
            (4,"Congés payés","2026-05-01","2026-05-08","Voyage familial","Approuvé",""),
            (8,"Maladie","2026-03-20","2026-03-22","", "Approuvé",""),
            (9,"Congés payés","2026-04-14","2026-04-18","", "Refusé","Planning chargé sur cette période, merci de proposer d'autres dates.")
        ])

        c.executemany("""
            INSERT INTO absences (emp_id,date,type,justifiee,note,duree)
            VALUES (?,?,?,?,?,?)
        """, [
            (4,"2026-03-15","Retard",1,"Transport",15),
            (9,"2026-03-18","Absence",0,"",0),
            (6,"2026-03-24","Retard",1,"RDV médical",30),
        ])

    # Forcer changement de mot de passe pour tous les employés créés en bulk
    c.execute("UPDATE users SET must_change_pw=1 WHERE role='employe'")

    conn.commit()
    conn.close()


# ── Auth helpers ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if "user_id" not in session:
            return jsonify({"error":"Non authentifié"}), 401
        return f(*a, **kw)
    return d


def role_required(*roles):
    def dec(f):
        @wraps(f)
        def d(*a, **kw):
            if "user_id" not in session:
                return jsonify({"error":"Non authentifié"}), 401
            if session.get("role") not in roles:
                return jsonify({"error":"Accès refusé"}), 403
            return f(*a, **kw)
        return d
    return dec


def current_user():
    return {
        "id": session.get("user_id"),
        "login": session.get("login"),
        "role": session.get("role"),
        "emp_id": session.get("emp_id"),
        "must_change_pw": session.get("must_change_pw", 0)
    }


# ── Pages ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if "user_id" not in session:
        return redirect("/login")
    return render_template("index.html", user=current_user())


@app.route("/login")
def login_page():
    return render_template("login.html")


# ── Auth API ────────────────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def do_login():
    d = request.json or {}
    login = d.get("login", "").strip()
    pw = d.get("password", "")
    ip = request.remote_addr or "?"
    key = f"{ip}|{login.lower()}"

    minutes = _login_check(key)
    if minutes:
        return jsonify({"error": f"Trop de tentatives. Réessayez dans {minutes} min."}), 429

    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE login=?", (login,)).fetchone()

    if not u or not verify_pw(u["password"], pw):
        conn.close()
        n = _login_fail(key)
        audit("Connexion échouée", f"login « {login} » · IP {ip} · tentative {n}", login="—")
        if n >= LOGIN_MAX_ATTEMPTS:
            audit("Connexions bloquées 15 min", f"login « {login} » · IP {ip}", login="—")
        return jsonify({"error":"Identifiants incorrects"}), 401

    # Migration transparente : re-hache les anciens mots de passe SHA-256 vers PBKDF2
    if not u["password"].startswith(("pbkdf2:", "scrypt:")):
        conn.execute("UPDATE users SET password=? WHERE id=?", (make_pw_hash(pw), u["id"]))
        conn.commit()
    conn.close()

    _login_attempts.pop(key, None)
    session.permanent = True
    session.update(
        user_id=u["id"],
        login=u["login"],
        role=u["role"],
        emp_id=u["emp_id"],
        must_change_pw=u["must_change_pw"]
    )
    audit("Connexion réussie", f"IP {ip}")
    return jsonify({
        "role": u["role"],
        "emp_id": u["emp_id"],
        "login": u["login"],
        "must_change_pw": bool(u["must_change_pw"])
    })


@app.route("/api/logout", methods=["POST"])
def do_logout():
    session.clear()
    return jsonify({"ok":True})


@app.route("/api/me")
@login_required
def me():
    conn = get_db()
    emp = None
    if session.get("emp_id"):
        row = conn.execute(
            "SELECT * FROM employees WHERE id=?",
            (session["emp_id"],)
        ).fetchone()
        emp = dict(row) if row else None
    conn.close()
    return jsonify({
        "current_user": current_user(),
        "employee": emp
    })


# ── Employees ───────────────────────────────────────────────────────────────

@app.route("/api/employees")
@login_required
def list_employees():
    conn = get_db()
    if session.get("role") == "employe":
        rows = conn.execute(
            "SELECT * FROM employees WHERE id=? AND actif=1",
            (session["emp_id"],)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM employees WHERE actif=1 ORDER BY nom,prenom"
        ).fetchall()
    result = []
    for r in rows:
        e = dict(r)
        e["anciennete"] = _anciennete(e.get("debut"))
        e["age"] = _age(e.get("date_naissance"))
        result.append(e)
    conn.close()
    return jsonify(result)


@app.route("/api/employees", methods=["POST"])
@role_required("directeur","rh")
def add_employee():
    d = request.json or {}
    if not d.get("prenom") or not d.get("nom"):
        return jsonify({"error":"Prénom et nom requis"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO employees (prenom,nom,poste,contrat,email,tel,tel_urgence,debut,date_fin,date_naissance,solde_cp)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        d["prenom"], d["nom"],
        d.get("poste",""),
        d.get("contrat","CDI"),
        d.get("email",""),
        d.get("tel",""),
        d.get("tel_urgence",""),
        d.get("debut",""),
        d.get("date_fin") or None,
        d.get("date_naissance") or None,
        int(d.get("solde_cp",25))
    ))
    conn.commit()
    row = conn.execute("SELECT * FROM employees WHERE id=?", (cur.lastrowid,)).fetchone()
    e = dict(row)
    e["anciennete"] = _anciennete(e.get("debut"))
    e["age"] = _age(e.get("date_naissance"))
    conn.close()
    return jsonify(e), 201


@app.route("/api/employees/<int:eid>")
@login_required
def get_employee(eid):
    if session.get("role") == "employe" and session.get("emp_id") != eid:
        return jsonify({"error":"Accès refusé"}), 403
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM employees WHERE id=?",
        (eid,)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"error":"Introuvable"}), 404
    emp = dict(row)
    emp["conges"] = [
        _conge_row(r) for r in conn.execute(
            "SELECT * FROM conges WHERE emp_id=? ORDER BY debut DESC",
            (eid,)
        ).fetchall()
    ]
    emp["absences"] = [
        dict(r) for r in conn.execute(
            "SELECT * FROM absences WHERE emp_id=? ORDER BY date DESC",
            (eid,)
        ).fetchall()
    ]
    conn.close()
    return jsonify(emp)


@app.route("/api/employees/<int:eid>", methods=["PUT"])
@login_required
def update_employee(eid):
    d = request.json or {}
    conn = get_db()
    if session.get("role") == "employe":
        if session.get("emp_id") != eid:
            conn.close()
            return jsonify({"error":"Accès refusé"}), 403
        conn.execute(
            "UPDATE employees SET email=?,tel=? WHERE id=?",
            (d.get("email",""), d.get("tel",""), eid)
        )
    else:
        conn.execute("""
            UPDATE employees SET prenom=?,nom=?,poste=?,contrat=?,email=?,tel=?,debut=?,date_fin=?,solde_cp=?
            WHERE id=?
        """, (
            d.get("prenom",""),
            d.get("nom",""),
            d.get("poste",""),
            d.get("contrat","CDI"),
            d.get("email",""),
            d.get("tel",""),
            d.get("debut",""),
            d.get("date_fin") or None,
            int(d.get("solde_cp",25)),
            eid
        ))
    conn.commit()
    row = conn.execute("SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route("/api/employees/<int:eid>", methods=["DELETE"])
@role_required("directeur")
def delete_employee(eid):
    conn = get_db()
    conn.execute("UPDATE employees SET actif=0 WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return jsonify({"ok":True})


# ── Congés ───────────────────────────────────────────────────────────────────

@app.route("/api/conges")
@login_required
def list_conges():
    conn = get_db()
    if session.get("role") == "employe":
        rows = conn.execute("""
            SELECT c.*, e.prenom||' '||e.nom AS emp_nom
            FROM conges c JOIN employees e ON e.id=c.emp_id
            WHERE c.emp_id=?
            ORDER BY c.id DESC
        """, (session["emp_id"],)).fetchall()
    else:
        rows = conn.execute("""
            SELECT c.*, e.prenom||' '||e.nom AS emp_nom
            FROM conges c JOIN employees e ON e.id=c.emp_id
            ORDER BY c.id DESC
        """).fetchall()
    conn.close()
    return jsonify([_conge_row(r) for r in rows])


@app.route("/api/conges", methods=["POST"])
@login_required
def add_conge():
    d = request.json or {}
    emp_id = session.get("emp_id") if session.get("role") == "employe" else d.get("emp_id")
    if not emp_id or not d.get("debut") or not d.get("fin"):
        return jsonify({"error":"Champs requis manquants"}), 400

    # Validation des dates
    try:
        d0 = datetime.strptime(d["debut"], "%Y-%m-%d").date()
        d1 = datetime.strptime(d["fin"], "%Y-%m-%d").date()
    except Exception:
        return jsonify({"error": "Format de date invalide"}), 400
    if d1 < d0:
        return jsonify({"error": "La date de fin doit être postérieure ou égale à la date de début"}), 400
    if jours_ouvres(d["debut"], d["fin"], d.get("creneau", "journee")) <= 0:
        return jsonify({"error": "La période ne contient aucun jour ouvré (week-end ou jour férié)"}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO conges (emp_id,type,debut,fin,creneau,motif,statut)
        VALUES (?,?,?,?,?, ?, 'En attente')
    """, (
        emp_id,
        d.get("type","Congés payés"),
        d["debut"],
        d["fin"],
        d.get("creneau","journee"),
        d.get("motif","").strip()
    ))
    conn.commit()
    row = conn.execute("""
        SELECT c.*, e.prenom||' '||e.nom AS emp_nom
        FROM conges c JOIN employees e ON e.id=c.emp_id
        WHERE c.id=?
    """, (cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify(_conge_row(row)), 201


@app.route("/api/conges/<int:cid>", methods=["PUT"])
@role_required("directeur","rh")
def update_conge(cid):
    d = request.json or {}
    new_statut = d.get("statut", "En attente")
    if new_statut not in ("En attente", "Approuvé", "Refusé"):
        return jsonify({"error": "Statut invalide"}), 400

    conn = get_db()
    cg = conn.execute("SELECT * FROM conges WHERE id=?", (cid,)).fetchone()
    if not cg:
        conn.close()
        return jsonify({"error": "Demande introuvable"}), 404

    old_statut = cg["statut"]
    days = jours_ouvres(cg["debut"], cg["fin"], cg["creneau"] or "journee")
    decompte = cg["type"] in TYPES_DECOMPTES  # seuls les congés payés touchent le solde

    # Passage vers Approuvé : contrôle du solde puis déduction
    if decompte and old_statut != "Approuvé" and new_statut == "Approuvé":
        row = conn.execute(
            "SELECT solde_cp FROM employees WHERE id=?", (cg["emp_id"],)
        ).fetchone()
        solde = (row["solde_cp"] or 0) if row else 0
        if days > solde:
            conn.close()
            return jsonify({
                "error": f"Solde insuffisant : {solde} j restants, demande de {days} j ouvrés"
            }), 409
        conn.execute(
            "UPDATE employees SET solde_cp = solde_cp - ? WHERE id=?",
            (days, cg["emp_id"])
        )

    # Sortie du statut Approuvé (refus ou remise en attente) : re-crédit exact
    if decompte and old_statut == "Approuvé" and new_statut != "Approuvé":
        conn.execute(
            "UPDATE employees SET solde_cp = solde_cp + ? WHERE id=?",
            (days, cg["emp_id"])
        )

    conn.execute(
        "UPDATE conges SET statut=?, note_refus=? WHERE id=?",
        (new_statut, d.get("note_refus", ""), cid)
    )
    conn.commit()
    row = conn.execute("""
        SELECT c.*, e.prenom||' '||e.nom AS emp_nom
        FROM conges c JOIN employees e ON e.id=c.emp_id
        WHERE c.id=?
    """, (cid,)).fetchone()
    conn.close()
    if new_statut != old_statut and new_statut in ("Approuvé", "Refusé"):
        audit(
            "Congé approuvé" if new_statut == "Approuvé" else "Congé refusé",
            f"{row['emp_nom']} · {row['type']} {row['debut']} → {row['fin']} · {days} j ouvrés"
        )
    return jsonify(_conge_row(row))


# ── Absences ────────────────────────────────────────────────────────────────

@app.route("/api/absences")
@login_required
def list_absences():
    conn = get_db()
    if session.get("role") == "employe":
        rows = conn.execute("""
            SELECT a.*, e.prenom||' '||e.nom AS emp_nom
            FROM absences a JOIN employees e ON e.id=a.emp_id
            WHERE a.emp_id=?
            ORDER BY a.date DESC, a.id DESC
        """, (session["emp_id"],)).fetchall()
    else:
        rows = conn.execute("""
            SELECT a.*, e.prenom||' '||e.nom AS emp_nom
            FROM absences a JOIN employees e ON e.id=a.emp_id
            ORDER BY a.date DESC, a.id DESC
        """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/absences", methods=["POST"])
@role_required("directeur","rh")
def add_absence():
    d = request.json or {}
    if not d.get("emp_id") or not d.get("date"):
        return jsonify({"error":"Champs requis manquants"}), 400

    emp_id = d.get("emp_id")
    date_str = d.get("date")
    typ = d.get("type","Absence")
    justifiee = 1 if d.get("justifiee") else 0
    note = d.get("note","").strip()

    # Durée en minutes pour retards
    duree = 0
    try:
        if typ == "Retard":
            duree = int(d.get("duree",0) or 0)
            if duree < 0:
                duree = 0
    except Exception:
        duree = 0

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO absences (emp_id,date,type,justifiee,note,duree)
        VALUES (?,?,?,?,?,?)
    """, (emp_id, date_str, typ, justifiee, note, duree))
    conn.commit()
    row = conn.execute("""
        SELECT a.*, e.prenom||' '||e.nom AS emp_nom
        FROM absences a JOIN employees e ON e.id=a.emp_id
        WHERE a.id=?
    """, (cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


# ── Dashboard ───────────────────────────────────────────────────────────────

@app.route("/api/dashboard")
@role_required("directeur","rh")
def dashboard():
    conn = get_db()

    today = date.today().isoformat()
    in_60 = (date.today() + timedelta(days=60)).isoformat()

    # Stats globales
    stats = {
        "total_emp": conn.execute("SELECT COUNT(*) FROM employees WHERE actif=1").fetchone()[0],
        "cdi": conn.execute("SELECT COUNT(*) FROM employees WHERE actif=1 AND contrat='CDI'").fetchone()[0],
        "cdd": conn.execute("SELECT COUNT(*) FROM employees WHERE actif=1 AND contrat='CDD'").fetchone()[0],
        "alt": conn.execute("SELECT COUNT(*) FROM employees WHERE actif=1 AND contrat='Alternant'").fetchone()[0],
        "ae": conn.execute("SELECT COUNT(*) FROM employees WHERE actif=1 AND contrat='Autoentrepreneur'").fetchone()[0],
        "att": conn.execute("SELECT COUNT(*) FROM conges WHERE statut='En attente'").fetchone()[0],
        "abs_mois": conn.execute(
            "SELECT COUNT(*) FROM absences WHERE strftime('%Y-%m',date)=strftime('%Y-%m','now')"
        ).fetchone()[0],
        "conges_mois": conn.execute(
            "SELECT COUNT(*) FROM conges WHERE statut='Approuvé' AND strftime('%Y-%m',debut)=strftime('%Y-%m','now')"
        ).fetchone()[0],
    }

    # Ancienneté
    rows = conn.execute(
        "SELECT debut FROM employees WHERE actif=1 AND debut IS NOT NULL"
    ).fetchall()
    total_years = 0.0
    count = 0
    five_plus = 0
    today_date = date.today()
    for r in rows:
        try:
            d0 = datetime.strptime(r["debut"], "%Y-%m-%d").date()
        except Exception:
            continue
        if d0 > today_date:
            continue
        delta_days = (today_date - d0).days
        years = delta_days / 365.0
        total_years += years
        count += 1
        if years >= 5.0:
            five_plus += 1
    avg_years = total_years / count if count else 0.0
    pct_five_plus = (five_plus / count * 100.0) if count else 0.0
    stats["avg_seniority_years"] = round(avg_years, 1)
    stats["pct_five_plus"] = round(pct_five_plus, 1)

    # Congés en attente
    pending = [
        _conge_row(r) for r in conn.execute("""
            SELECT c.*, e.prenom||' '||e.nom AS emp_nom
            FROM conges c JOIN employees e ON e.id=c.emp_id
            WHERE c.statut='En attente'
            ORDER BY c.created_at
        """).fetchall()
    ]

    # Fins de contrat dans 60 jours
    expiring = [
        dict(r) for r in conn.execute("""
            SELECT id, prenom, nom, contrat, poste, date_fin
            FROM employees
            WHERE actif=1 AND date_fin IS NOT NULL
              AND date_fin >= ? AND date_fin <= ?
            ORDER BY date_fin
        """, (today, in_60)).fetchall()
    ]

    # Absences récentes 7 jours
    recent_abs = [
        dict(r) for r in conn.execute("""
            SELECT a.*, e.prenom||' '||e.nom AS emp_nom
            FROM absences a JOIN employees e ON e.id=a.emp_id
            WHERE a.date >= date('now','-7 days')
            ORDER BY a.date DESC
        """).fetchall()
    ]

    # Congés par mois 6 derniers mois
    monthly = [
        dict(r) for r in conn.execute("""
            SELECT strftime('%Y-%m', debut) AS mois,
                   COUNT(*) AS total,
                   SUM(CASE WHEN statut='Approuvé' THEN 1 ELSE 0 END) AS approuves
            FROM conges
            WHERE debut >= date('now','-6 months')
            GROUP BY mois
            ORDER BY mois
        """).fetchall()
    ]

    conn.close()
    return jsonify({
        "stats": stats,
        "pending": pending,
        "expiring": expiring,
        "recent_abs": recent_abs,
        "monthly": monthly
    })


# ── Users ───────────────────────────────────────────────────────────────────

@app.route("/api/users")
@role_required("directeur")
def list_users():
    conn = get_db()
    rows = conn.execute("""
        SELECT u.id, u.login, u.role, u.emp_id, u.must_change_pw,
               e.prenom||' '||e.nom AS emp_nom
        FROM users u LEFT JOIN employees e ON e.id=u.emp_id
        ORDER BY u.role, u.login
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/users", methods=["POST"])
@role_required("directeur")
def add_user():
    d = request.json or {}
    if not d.get("login") or not d.get("password"):
        return jsonify({"error":"Login et mot de passe requis"}), 400
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO users (login,password,role,emp_id,must_change_pw)
            VALUES (?,?,?,?,1)
        """, (
            d["login"],
            make_pw_hash(d["password"]),
            d.get("role","employe"),
            d.get("emp_id") or None
        ))
        conn.commit()
    except Exception:
        conn.close()
        return jsonify({"error":"Login déjà existant"}), 400
    conn.close()
    audit("Compte créé", f"login « {d['login']} » · rôle {d.get('role','employe')}")
    return jsonify({"ok":True}), 201


@app.route("/api/users/<int:uid>", methods=["PUT"])
@role_required("directeur")
def update_user(uid):
    d = request.json or {}
    conn = get_db()
    target = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return jsonify({"error":"Compte introuvable"}), 404
    new_role = d.get("role","employe")
    # Protection : ne pas rétrograder le dernier compte directeur
    if target["role"] == "directeur" and new_role != "directeur":
        nb = conn.execute("SELECT COUNT(*) FROM users WHERE role='directeur'").fetchone()[0]
        if nb <= 1:
            conn.close()
            return jsonify({"error":"Impossible de retirer le rôle du dernier compte directeur"}), 400
    if d.get("password"):
        conn.execute("""
            UPDATE users SET role=?, emp_id=?, password=?, must_change_pw=1
            WHERE id=?
        """, (
            new_role, d.get("emp_id") or None,
            make_pw_hash(d["password"]), uid
        ))
    else:
        conn.execute("""
            UPDATE users SET role=?, emp_id=?
            WHERE id=?
        """, (
            new_role, d.get("emp_id") or None, uid
        ))
    conn.commit()
    conn.close()
    audit("Compte modifié",
          f"login « {target['login']} » · rôle {new_role}"
          + (" · mot de passe réinitialisé" if d.get("password") else ""))
    return jsonify({"ok":True})


@app.route("/api/users/<int:uid>", methods=["DELETE"])
@role_required("directeur")
def delete_user(uid):
    conn = get_db()
    target = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return jsonify({"error":"Compte introuvable"}), 404
    if target["id"] == session.get("user_id"):
        conn.close()
        return jsonify({"error":"Impossible de supprimer votre propre compte"}), 400
    if target["role"] == "directeur":
        nb = conn.execute("SELECT COUNT(*) FROM users WHERE role='directeur'").fetchone()[0]
        if nb <= 1:
            conn.close()
            return jsonify({"error":"Impossible de supprimer le dernier compte directeur"}), 400
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    audit("Compte supprimé", f"login « {target['login']} » · rôle {target['role']}")
    return jsonify({"ok":True})


@app.route("/api/change-password", methods=["POST"])
@login_required
def change_password():
    d = request.json or {}
    if not d.get("new_password") or len(d["new_password"]) < 8:
        return jsonify({"error":"Le nouveau mot de passe doit contenir au moins 8 caractères"}), 400
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
    if not u or not verify_pw(u["password"], d.get("old_password","")):
        conn.close()
        return jsonify({"error":"Ancien mot de passe incorrect"}), 400
    conn.execute(
        "UPDATE users SET password=?, must_change_pw=0 WHERE id=?",
        (make_pw_hash(d["new_password"]), session["user_id"])
    )
    conn.commit()
    conn.close()
    session["must_change_pw"] = 0
    audit("Mot de passe modifié", "changement par l'utilisateur")
    return jsonify({"ok":True})


# ── Sauvegarde manuelle (bouton Directeur) ──────────────────────────────────

@app.route("/api/backup", methods=["POST"])
@role_required("directeur")
def api_backup():
    name = backup_db()
    if not name:
        return jsonify({"error": "Sauvegarde impossible (base introuvable ou erreur d'écriture)"}), 500
    audit("Sauvegarde de la base", name)
    return jsonify({"ok": True, "fichier": name})


# ── Journal d'audit (consultation) ──────────────────────────────────────────

@app.route("/api/audit")
@role_required("directeur")
def list_audit():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM audit_log ORDER BY id DESC LIMIT 500"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ── Export / Import Excel ───────────────────────────────────────────────────

@app.route("/api/export")
@role_required("directeur","rh")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error":"openpyxl non installé"}), 500

    conn = get_db()
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfil = PatternFill("solid", fgColor="16A37F")
    ctr = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin,right=thin,top=thin,bottom=thin)

    def sh(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hf
            cell.fill = hfil
            cell.alignment = ctr
            cell.border = brd

    def sr(ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = brd
                cell.alignment = Alignment(vertical="center")

    # Annuaire
    ws1 = wb.active
    ws1.title = "Annuaire RH"
    sh(ws1, ["ID","Prénom","Nom","Poste","Contrat","Email","Téléphone","Entrée","Fin contrat","Solde CP"])
    for r in conn.execute("""
        SELECT id,prenom,nom,poste,contrat,email,tel,debut,date_fin,solde_cp
        FROM employees WHERE actif=1 ORDER BY nom
    """).fetchall():
        ws1.append(list(r))
    sr(ws1)
    for col,w in zip("ABCDEFGHIJ",[6,12,15,22,15,28,14,12,12,9]):
        ws1.column_dimensions[col].width = w

    # Congés
    ws2 = wb.create_sheet("Congés")
    sh(ws2, ["ID","Collaborateur","Type","Créneau","Du","Au","Jours ouvrés","Motif","Statut","Note refus","Soumis le"])
    for r in conn.execute("""
        SELECT c.id, e.prenom||' '||e.nom, c.type, c.creneau,
               c.debut, c.fin,
               c.motif, c.statut, c.note_refus, c.created_at
        FROM conges c JOIN employees e ON e.id=c.emp_id
        ORDER BY c.id DESC
    """).fetchall():
        vals = list(r)
        jo = jours_ouvres(vals[4], vals[5], vals[3] or "journee")
        ws2.append(vals[:6] + [jo] + vals[6:])
    sr(ws2)
    for col,w in zip("ABCDEFGHIJK",[6,20,16,10,12,12,7,20,12,24,16]):
        ws2.column_dimensions[col].width = w

    # Absences
    ws3 = wb.create_sheet("Absences")
    sh(ws3, ["ID","Collaborateur","Date","Type","Durée (min)","Justifiée","Note","Enregistré le"])
    for r in conn.execute("""
        SELECT a.id, e.prenom||' '||e.nom, a.date, a.type,
               COALESCE(a.duree,0),
               CASE WHEN a.justifiee=1 THEN 'Oui' ELSE 'Non' END,
               a.note, a.created_at
        FROM absences a JOIN employees e ON e.id=a.emp_id
        ORDER BY a.date DESC
    """).fetchall():
        ws3.append(list(r))
    sr(ws3)
    for col,w in zip("ABCDEFGH",[6,20,12,10,12,10,20,16]):
        ws3.column_dimensions[col].width = w

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"epie_formation_rh_{datetime.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/import-employees", methods=["POST"])
@role_required("directeur","rh")
def import_employees():
    f = request.files.get("file")
    if not f:
        return jsonify({"error":"Fichier manquant (champ 'file')"}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error":"openpyxl non installé"}), 500

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({"error":f"Impossible de lire le fichier Excel: {e}"}), 400

    ws = wb.active
    conn = get_db()
    cur = conn.cursor()

    imported = 0
    updated = 0
    skipped = 0

    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not row:
            continue
        if len(row) < 10:
            skipped += 1
            continue

        _id, prenom, nom, poste, contrat, email, tel, debut, date_fin, solde_cp = row
        if not prenom or not nom:
            skipped += 1
            continue

        email = email or ""
        tel = tel or ""
        poste = poste or ""
        contrat = contrat or "CDI"
        if hasattr(debut, "strftime"):
            debut = debut.strftime("%Y-%m-%d")
        else:
            debut = debut or ""
        if hasattr(date_fin, "strftime"):
            date_fin = date_fin.strftime("%Y-%m-%d")
        else:
            date_fin = date_fin or None
        try:
            solde_cp = int(solde_cp or 25)
        except Exception:
            solde_cp = 25

        if email:
            existing = conn.execute(
                "SELECT id FROM employees WHERE email=? AND actif=1",
                (email,)
            ).fetchone()
        else:
            existing = None

        if existing:
            emp_id = existing["id"]
            cur.execute("""
                UPDATE employees
                SET prenom=?, nom=?, poste=?, contrat=?, tel=?, debut=?, date_fin=?, solde_cp=?
                WHERE id=?
            """, (prenom, nom, poste, contrat, tel, debut or None, date_fin, solde_cp, emp_id))
            updated += 1
        else:
            cur.execute("""
                INSERT INTO employees (prenom,nom,poste,contrat,email,tel,debut,date_fin,solde_cp,actif)
                VALUES (?,?,?,?,?,?,?,?,?,1)
            """, (prenom, nom, poste, contrat, email, tel, debut or None, date_fin, solde_cp))
            imported += 1

    conn.commit()
    conn.close()
    return jsonify({"ok":True, "imported":imported, "updated":updated, "skipped":skipped})


@app.route("/api/import-conges-absences", methods=["POST"])
@role_required("directeur","rh")
def import_conges_absences():
    f = request.files.get("file")
    if not f:
        return jsonify({"error":"Fichier manquant (champ 'file')"}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error":"openpyxl non installé"}), 500

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({"error":f"Impossible de lire le fichier Excel: {e}"}), 400

    conn = get_db()
    cur = conn.cursor()

    imported_conges = 0
    imported_abs = 0
    skipped_conges = 0
    skipped_abs = 0

    def get_emp_id(full_name):
        if not full_name:
            return None
        parts = str(full_name).strip().split()
        if len(parts) < 2:
            return None
        prenom = parts[0]
        nom = " ".join(parts[1:])
        row = conn.execute(
            "SELECT id FROM employees WHERE prenom=? AND nom=? AND actif=1",
            (prenom, nom)
        ).fetchone()
        return row["id"] if row else None

    # Congés
    if "Congés" in wb.sheetnames:
        ws_c = wb["Congés"]
        first = True
        for row in ws_c.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row:
                continue
            if len(row) < 11:
                skipped_conges += 1
                continue
            _id, collab, typ, creneau, debut, fin, _jours, motif, statut, note_refus, created_at = row
            emp_id = get_emp_id(collab)
            if not emp_id or not debut or not fin:
                skipped_conges += 1
                continue
            if hasattr(debut, "strftime"):
                debut = debut.strftime("%Y-%m-%d")
            else:
                debut = str(debut)
            if hasattr(fin, "strftime"):
                fin = fin.strftime("%Y-%m-%d")
            else:
                fin = str(fin)
            if hasattr(created_at, "strftime"):
                created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
            else:
                created_at = None
            typ = typ or "Congés payés"
            creneau = creneau or "journee"
            motif = motif or ""
            statut = statut or "En attente"
            note_refus = note_refus or ""
            cur.execute("""
                INSERT INTO conges (emp_id,type,debut,fin,creneau,motif,statut,note_refus,created_at)
                VALUES (?,?,?,?,?,?,?, ?, COALESCE(?, datetime('now','localtime')))
            """, (emp_id, typ, debut, fin, creneau, motif, statut, note_refus, created_at))
            imported_conges += 1

    # Absences
    if "Absences" in wb.sheetnames:
        ws_a = wb["Absences"]
        first = True
        for row in ws_a.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row:
                continue
            if len(row) < 8:
                skipped_abs += 1
                continue
            _id, collab, date_abs, typ, duree, justif_str, note, created_at = row
            emp_id = get_emp_id(collab)
            if not emp_id or not date_abs:
                skipped_abs += 1
                continue
            if hasattr(date_abs, "strftime"):
                date_abs = date_abs.strftime("%Y-%m-%d")
            else:
                date_abs = str(date_abs)
            if hasattr(created_at, "strftime"):
                created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
            else:
                created_at = None
            typ = typ or "Absence"
            note = note or ""
            justifiee = 1 if str(justif_str).strip().lower().startswith("oui") else 0
            try:
                duree_val = int(duree or 0)
            except Exception:
                duree_val = 0
            cur.execute("""
                INSERT INTO absences (emp_id,date,type,justifiee,note,duree,created_at)
                VALUES (?,?,?,?,?,?, COALESCE(?, datetime('now','localtime')))
            """, (emp_id, date_abs, typ, justifiee, note, duree_val, created_at))
            imported_abs += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "imported_conges": imported_conges,
        "skipped_conges": skipped_conges,
        "imported_absences": imported_abs,
        "skipped_absences": skipped_abs
    })


if __name__ == "__main__":
    init_db()
    backup_db()  # sauvegarde automatique à chaque démarrage
    app.run(host="0.0.0.0", port=5000, debug=False)